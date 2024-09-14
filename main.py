#%% imports

from libs import *


import ConfigSetup
import forum_trans
import forum_api
import pandas as pd
import openpyxl

##Adjust file writing to be incremental line by line or convert to db

#%% 
ConfigSetup.initial_setup()
config = ConfigSetup.read_cfg()


after_time = datetime(2021, 1, 1, 00, 00, 00)
from_unix = int(time.mktime(after_time.timetuple()))

before_time = datetime(2021, 6, 13, 23, 59, 59)
to_unix = int(time.mktime(before_time.timetuple()))


#%% Get Threads and Posts

# for category in cats:
#     if category["id"] != 2:
#         continue
#     print(category["title"])
    
result = forum_api.get_forum_threads(config['main']['primary_api'], from_unix, to_unix)
#%%  Post File Generation

with open(str(config['main']['post_file']) + ".json", 'r') as f:
    pst_obj = json.load(f)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Post Data"
headers = [
    "ID", "Thread ID", "Is Legacy", "Is Topic", "Is Edited", "Is Pinned", "Created Time",
    "Edited By", "Has Quote", "Quoted Post ID", "Content", "Likes", "Dislikes", "Forum ID",
    "Pos", "Neg", "Neu", "Compound", "User ID", "Username", "Karma","Length"
]
ws.append(headers)
for post_id, post_info in pst_obj.items():
    try:
        row = [
            post_info.get("id", ""),
            post_info.get("thread_id", None),
            post_info.get("is_legacy", None),
            post_info.get("is_topic", None),
            post_info.get("is_edited", None),
            post_info.get("is_pinned", None),
            post_info.get("created_time", None),
            post_info.get("edited_by", None),
            post_info.get("has_quote", None),
            post_info.get("quoted_post_id", ""),
            post_info.get("content", ""),
            post_info.get("likes", None),
            post_info.get("dislikes", None),
            post_info.get("forum_id", None),
            post_info.get("pos", None),
            post_info.get("neg", None),
            post_info.get("neu", None),
            post_info.get("compound", None),
            post_info.get("user_id", None),
            post_info.get("username", ""),
            post_info.get("karma", None),
            post_info.get("length", None)
        ]
        ws.append(row)
    except:
        print("Failed: " + str(post_info.get("id", ""))) #Some rows have errors from Torn

# Save the Excel file - Excel is necessary for Tableau
wb.save(str(config['main']['post_file']) + ".xlsx")

#%%  Thread File Generation

with open(str(config['main']['thread_file']) + ".json", 'r') as f:
    thd_obj = json.load(f)
    
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Thread Data"

# Define the headers
headers = [
    "ID", "Title", "Forum ID", "Posts", "Rating", "Views", "First Post Time",
    "Last Post Time", "Has Poll", "Is Locked", "Is Sticky", "Content",
    "Pos", "Neg", "Neu", "Compound", "User ID", "Username", "Karma"
]
ws.append(headers)
for thread_id, thread_data in thd_obj.items():
    try:
        row = [
            thread_data.get("id", None),
            thread_data.get("title", ""),
            thread_data.get("forum_id", None),
            thread_data.get("posts", None),
            thread_data.get("rating", None),
            thread_data.get("views", None),
            thread_data.get("first_post_time", None),
            thread_data.get("last_post_time", None),
            thread_data.get("has_poll", None),
            thread_data.get("is_locked", None),
            thread_data.get("is_sticky", None),
            thread_data.get("content", ""),
            thread_data.get("pos", None),
            thread_data.get("neg", None),
            thread_data.get("neu", None),
            thread_data.get("compound", None),
            thread_data.get("user_id", None),
            thread_data.get("username", ""),
            thread_data.get("karma", None)
        ]
        ws.append(row)
    except:
        print("Failed: " + str(thread_data.get("id", None)))
              
# Save the Excel file
wb.save(str(config['main']['thread_file']) + ".xlsx")

#%% User File Generation

with open(str(config['main']['user_file']) + ".json", 'r') as f:
    usr_obj = json.load(f)
table_data = [
    {"Username": outer_key, "Word": inner_key, "Value": inner_value}
    for outer_key, inner_dict in usr_obj.items()
    for inner_key, inner_value in inner_dict.items()
]



wb = openpyxl.Workbook()
ws = wb.active
ws.title = "User Data"
ws.append(["Username", "Word", "Value"])
for row in table_data:
    try:
        ws.append([row["Username"], row["Word"], row["Value"]])
    except:
        print("Unprintable: " + str(row["Word"]))
wb.save(str(config['main']['user_file']) + ".xlsx")


#%%   Category File Generation
cats = forum_api.get_forum_categories(config['main']['primary_api'])
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Category Data"
headers = ["ID", "Title", "Acronym", "Threads"]
ws.append(headers)
for cat in cats:
    row = [
        cat.get("id", None),
        cat.get("title", ""),
        cat.get("acronym", ""),
        cat.get("threads", None)
    ]
    ws.append(row)

# Save the Excel file
wb.save("category_data.xlsx")



